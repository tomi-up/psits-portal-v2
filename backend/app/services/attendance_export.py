"""Builds the Excel (.xlsx) attendance report for an event.

Renders exactly the rows/statuses produced by
app.api.v1.endpoints.admin_events.build_event_registrations - never queries
the database itself - so the DataTable and the export can't drift apart.
"""

from __future__ import annotations

import re
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

if TYPE_CHECKING:
    from app.api.v1.endpoints.admin_events import EventRegistrationsResponse

ORG_NAME = "PHILIPPINE SOCIETY OF INFORMATION TECHNOLOGY STUDENTS"
PH_UTC_OFFSET = timedelta(hours=8)

# psits-portal-v2/backend/app/services/attendance_export.py -> psits-portal-v2/frontend/public/psits-logo.png
LOGO_PATH = Path(__file__).resolve().parents[3] / "frontend" / "public" / "psits-logo.png"

REGISTRATION_LABELS = {
    "REGISTERED": "Registered",
    "NOT_REGISTERED": "Not Registered",
    "NOT_REQUIRED": "Not Required",
}

STATUS_LABELS = {
    "PRESENT": "Present",
    "INCOMPLETE": "Incomplete",
    "NO_SHOW": "No Show",
    "ABSENT": "Absent",
    "NOT_REGISTERED": "Not Registered",
    "EXCUSED": "Excused",
}

# Rendered in this order in both the summary block and the status column,
# regardless of which ones actually occur for a given event.
STATUS_ORDER = ["PRESENT", "INCOMPLETE", "NO_SHOW", "ABSENT", "NOT_REGISTERED", "EXCUSED"]

COLUMNS = [
    ("Student ID", 16),
    ("Last Name", 20),
    ("First Name", 20),
    ("Middle Name", 18),
    ("Course", 10),
    ("Year", 10),
    ("Section", 10),
    ("Registration Status", 18),
    ("Attendance Status", 18),
    ("Time In", 12),
    ("Time Out", 12),
]

THIN_BORDER = Border(*(Side(style="thin", color="B0B7C3") for _ in range(4)))
HEADER_FILL = PatternFill("solid", fgColor="0F172A")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def safe_filename(name: str) -> str:
    """Strip characters that break on common filesystems and collapse whitespace."""
    stripped = re.sub(r'[/\\:*?"<>|]', "", name)
    stripped = re.sub(r"\s+", "_", stripped.strip())
    return stripped or "PSITS_Attendance"


def _ordinal_year(year_level: int | None) -> str:
    if year_level is None:
        return "—"
    suffix = {1: "ST", 2: "ND", 3: "RD"}.get(year_level, "TH")
    return f"{year_level}{suffix}"


def _to_ph_time_str(dt: datetime | None) -> str:
    """time_in/time_out arrive as naive-but-UTC (see app/core/attendance.py);
    shift to PH local wall-clock for a report officers will read as-is."""
    if not dt:
        return "—"
    ph = dt.replace(tzinfo=None) + PH_UTC_OFFSET
    return ph.strftime("%I:%M %p").lstrip("0")


def _write_header(ws: Worksheet, data: "EventRegistrationsResponse") -> int:
    """Writes the logo/org/event header block, returns the next free row."""
    row = 1

    if LOGO_PATH.exists():
        img = XLImage(str(LOGO_PATH))
        img.width = 64
        img.height = 64
        ws.add_image(img, "A1")
        ws.row_dimensions[1].height = 48
        ws.row_dimensions[2].height = 48
        title_col = "B"
    else:
        title_col = "A"

    last_col = get_column_letter(len(COLUMNS))

    ws.merge_cells(f"{title_col}1:{last_col}1")
    ws[f"{title_col}1"] = ORG_NAME
    ws[f"{title_col}1"].font = Font(bold=True, size=13)

    ws.merge_cells(f"{title_col}2:{last_col}2")
    ws[f"{title_col}2"] = "ATTENDANCE REPORT"
    ws[f"{title_col}2"].font = Font(bold=True, size=11, color="475569")
    row = 3

    row += 1
    ws.merge_cells(f"A{row}:{last_col}{row}")
    ws[f"A{row}"] = data.event_name
    ws[f"A{row}"].font = Font(bold=True, size=14)

    row += 1
    ws.merge_cells(f"A{row}:{last_col}{row}")
    date_str = data.event_date.strftime("%B %d, %Y") if data.event_date else "Date not set"
    ws[f"A{row}"] = f"Event Date: {date_str}"
    ws[f"A{row}"].font = Font(size=10, color="475569")

    return row + 2


def _write_summary(ws: Worksheet, data: "EventRegistrationsResponse", start_row: int) -> int:
    row = start_row
    last_col = get_column_letter(len(COLUMNS))

    ws.merge_cells(f"A{row}:{last_col}{row}")
    ws[f"A{row}"] = "ATTENDANCE SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=11)
    row += 1

    counts = {
        "PRESENT": data.total_present,
        "INCOMPLETE": data.total_incomplete,
        "NO_SHOW": data.total_no_show,
        "ABSENT": data.total_absent,
        "NOT_REGISTERED": data.total_not_registered,
        "EXCUSED": data.total_excused,
    }
    total_eligible = len(data.registrations)

    summary_lines = [("Total Eligible Students", total_eligible)]
    for key in STATUS_ORDER:
        if counts[key] > 0:
            summary_lines.append((STATUS_LABELS[key], counts[key]))
    if data.total_late > 0:
        summary_lines.append(("Late Arrivals", data.total_late))

    for label, value in summary_lines:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=(label == "Total Eligible Students"))
        ws[f"B{row}"] = value
        ws[f"B{row}"].font = Font(bold=(label == "Total Eligible Students"))
        row += 1

    return row + 1


def _write_table(ws: Worksheet, data: "EventRegistrationsResponse", start_row: int) -> None:
    header_row = start_row
    for col_idx, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for offset, r in enumerate(data.registrations, start=1):
        row = header_row + offset
        values = [
            r.student_id,
            r.last_name,
            r.first_name,
            r.middle_name or "—",
            r.program or "—",
            _ordinal_year(r.year_level),
            r.section or "—",
            REGISTRATION_LABELS.get(r.registration_status, r.registration_status),
            STATUS_LABELS.get(r.status, r.status),
            _to_ph_time_str(r.time_in),
            _to_ph_time_str(r.time_out),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col_idx not in (2, 3, 4) else "left",
                vertical="center",
                wrap_text=col_idx in (2, 3, 4),
            )

    last_row = header_row + len(data.registrations)
    last_col_letter = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_row}"
    ws.freeze_panes = f"A{header_row + 1}"


def build_attendance_workbook(data: "EventRegistrationsResponse") -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    next_row = _write_header(ws, data)
    next_row = _write_summary(ws, data, next_row)
    _write_table(ws, data, next_row)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
