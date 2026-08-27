"""
PSITS V2 - Current Student Excel Importer (LOCAL PROCESSING ONLY)

Pipeline:
    Excel -> local processing -> normalization -> validation -> preview/report -> STOP

There is NO database write of any kind in this script. It never imports a
Supabase client, never reads Supabase credentials, and never touches the
main application's `students` / `profiles` tables. It only reads the source
workbook (read-only) and writes a local preview CSV + report under
private_data/import_preview/, which is gitignored.

Usage:
    python import_students.py [path/to/workbook.xlsx]

    Defaults to private_data/PSITS_ATTENDANCE_SY_2026-2027.xlsx (relative to
    the project root, two levels up from this script) if no path is given.
"""

import csv
import sys
from dataclasses import asdict
from pathlib import Path

import openpyxl

from normalize import normalize_student
from validate import validate_student, detect_duplicates, detect_conflicts, ValidatedStudent

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_WORKBOOK = PROJECT_ROOT / "private_data" / "PSITS_ATTENDANCE_SY_2026-2027.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "private_data" / "import_preview"

# Sheets we process, in report order. Matched case-insensitively as a
# substring against the actual workbook sheet names, since real workbooks
# have been observed to use "1ST YR" rather than the spec's "1st".
STUDENT_SHEET_MARKERS = ["1ST", "2ND", "3RD", "4TH"]
OFFICER_SHEET_MARKER = "OFFICER"

IDENTITY_COLUMNS = {
    "student_id": "STUDENT ID",
    "last_name": "LAST NAME",
    "first_name": "FIRST NAME",
    "middle_name": "MIDDLE NAME",
    "year_level": "YEAR",
    "course": "COURSE",
    "section": "SECTION",
}


def classify_sheets(sheet_names: list[str]) -> tuple[list[str], list[str], list[str]]:
    """Returns (student_sheets_in_order, officer_sheets, unrecognized_sheets)."""
    student_sheets: list[str | None] = [None] * len(STUDENT_SHEET_MARKERS)
    officer_sheets = []
    unrecognized = []

    for name in sheet_names:
        upper = name.upper()
        if OFFICER_SHEET_MARKER in upper:
            officer_sheets.append(name)
            continue
        matched = False
        for i, marker in enumerate(STUDENT_SHEET_MARKERS):
            if marker in upper:
                student_sheets[i] = name
                matched = True
                break
        if not matched:
            unrecognized.append(name)

    missing = [STUDENT_SHEET_MARKERS[i] for i, s in enumerate(student_sheets) if s is None]
    if missing:
        print(f"STOP: expected student-year sheets not found for: {missing}")
        print(f"Actual sheets in workbook: {sheet_names}")
        sys.exit(1)

    return [s for s in student_sheets if s], officer_sheets, unrecognized


def find_header_row(ws, max_scan_rows: int = 20) -> int | None:
    """Locate the row whose STUDENT ID column header matches, by scanning the
    first N rows rather than assuming a fixed row number."""
    for r in range(1, max_scan_rows + 1):
        for c in range(1, 15):
            value = ws.cell(row=r, column=c).value
            if value and str(value).strip().upper() == "STUDENT ID":
                return r
    return None


def map_columns(ws, header_row: int) -> dict[str, int] | None:
    """Map each identity field to its column index by matching header text -
    robust to column reordering, unlike hardcoded column letters."""
    header_values = {}
    for c in range(1, 30):
        value = ws.cell(row=header_row, column=c).value
        if value:
            header_values[str(value).strip().upper()] = c

    mapping = {}
    for field, header_label in IDENTITY_COLUMNS.items():
        col = header_values.get(header_label)
        if col is None:
            return None
        mapping[field] = col
    return mapping


def _blank(value) -> bool:
    return value is None or str(value).strip() == ""


def extract_student_records(ws, sheet_label: str) -> list[dict]:
    """A row counts as a data row iff STUDENT ID, LAST NAME, or FIRST NAME is
    non-blank - this naturally skips the title rows, the header row, the
    course-group header rows (e.g. "BACHELOR OF SCIENCE IN COMPUTER
    SCIENCE"), and blank separator rows that this workbook interleaves
    between course groups, without guessing based on SEQ. NO. formatting.

    Checking all three identity columns (not just STUDENT ID) matters: this
    workbook has real rows with a populated name but a BLANK Student ID cell
    (a data-entry gap) - using Student ID alone as the signal would silently
    drop those students instead of flagging them as invalid."""
    header_row = find_header_row(ws)
    if header_row is None:
        print(f"STOP: could not locate the 'STUDENT ID' header in sheet '{sheet_label}'.")
        sys.exit(1)

    columns = map_columns(ws, header_row)
    if columns is None:
        print(f"STOP: sheet '{sheet_label}' is missing one or more required identity columns.")
        print(f"Expected: {list(IDENTITY_COLUMNS.values())}")
        sys.exit(1)

    records = []
    max_row = ws.max_row or 0
    for r in range(header_row + 1, max_row + 1):
        student_id_cell = ws.cell(row=r, column=columns["student_id"]).value
        last_name_cell = ws.cell(row=r, column=columns["last_name"]).value
        first_name_cell = ws.cell(row=r, column=columns["first_name"]).value
        if _blank(student_id_cell) and _blank(last_name_cell) and _blank(first_name_cell):
            continue
        raw = {field: ws.cell(row=r, column=col).value for field, col in columns.items()}
        records.append({"raw": raw, "row": r})

    return records


def run(workbook_path: Path) -> None:
    if not workbook_path.exists():
        print(f"STOP: workbook not found at {workbook_path}")
        sys.exit(1)

    print("Loading workbook...")
    # NOTE: read_only=True was tried first, but this workbook's dimension
    # metadata is unreliable under it (ws.max_row comes back None), which
    # silently produced zero extracted rows. Loading normally instead - we
    # never call wb.save() anywhere, so the source file is still never
    # written back to; this only affects how openpyxl reads it into memory.
    wb = openpyxl.load_workbook(workbook_path, data_only=True)

    student_sheets, officer_sheets, unrecognized = classify_sheets(wb.sheetnames)

    print(f"Found sheets: {wb.sheetnames}")
    for name in officer_sheets:
        print(f"Found sheet: {name}")
        print("Status: SKIPPED - OUT OF SCOPE")
    if unrecognized:
        print(f"Found unrecognized sheet(s), skipped (not student or officer): {unrecognized}")
    print()

    per_sheet_validated: dict[str, list[ValidatedStudent]] = {}
    all_validated: list[ValidatedStudent] = []

    for sheet_name in student_sheets:
        ws = wb[sheet_name]
        raw_rows = extract_student_records(ws, sheet_name)

        validated_rows = []
        for item in raw_rows:
            normalized = normalize_student(item["raw"], source_sheet=sheet_name, source_row=item["row"])
            validated_rows.append(validate_student(normalized))

        per_sheet_validated[sheet_name] = validated_rows
        all_validated.extend(validated_rows)

    global_duplicates = detect_duplicates(all_validated)
    global_conflicts = detect_conflicts(global_duplicates)

    generate_validation_report(
        workbook_path, student_sheets, officer_sheets, per_sheet_validated, all_validated, global_duplicates, global_conflicts
    )
    generate_preview(all_validated)

    print_terminal_sample(all_validated)


def generate_validation_report(
    workbook_path: Path,
    student_sheets: list[str],
    officer_sheets: list[str],
    per_sheet_validated: dict[str, list[ValidatedStudent]],
    all_validated: list[ValidatedStudent],
    global_duplicates: dict[str, list[ValidatedStudent]],
    global_conflicts: dict[str, list[ValidatedStudent]],
) -> None:
    lines = []
    lines.append("PSITS CURRENT STUDENT IMPORT VALIDATION")
    lines.append("")
    lines.append("Workbook:")
    lines.append(str(workbook_path))
    lines.append("")
    lines.append("Processed sheets:")
    for s in student_sheets:
        lines.append(f"- {s}")
    lines.append("")
    lines.append("Skipped:")
    for s in officer_sheets:
        lines.append(f"- {s}")
    lines.append("")
    lines.append("-" * 32)

    total_source_rows = 0
    total_valid = 0
    total_invalid = 0
    total_missing_student_id = 0
    total_missing_fields = 0
    total_os = 0

    for sheet_name in student_sheets:
        rows = per_sheet_validated[sheet_name]
        sheet_dupes = detect_duplicates(rows)
        valid = sum(1 for v in rows if v.is_valid)
        invalid = sum(1 for v in rows if not v.is_valid)
        os_count = sum(1 for v in rows if v.is_over_stay)

        total_source_rows += len(rows)
        total_valid += valid
        total_invalid += invalid
        total_missing_student_id += sum(
            1 for v in rows if not v.record.student_id
        )
        total_missing_fields += sum(
            1 for v in rows for i in v.issues if i.code == "MISSING_REQUIRED_FIELD"
        )
        total_os += os_count

        lines.append("")
        lines.append(sheet_name)
        lines.append(f"Source rows: {len(rows)}")
        lines.append(f"Valid: {valid}")
        lines.append(f"Invalid: {invalid}")
        lines.append(f"Duplicates (within sheet): {len(sheet_dupes)}")
        if "4TH" in sheet_name.upper():
            lines.append(f"Over Stay (OS): {os_count}")
        else:
            lines.append(f"OS: {os_count}")

    lines.append("")
    lines.append("-" * 32)
    lines.append("")
    lines.append("TOTAL")
    lines.append("")
    lines.append(f"Source rows: {total_source_rows}")
    lines.append(f"Valid records: {total_valid}")
    lines.append(f"Invalid records: {total_invalid}")
    lines.append(f"Duplicate Student IDs: {len(global_duplicates)}")
    lines.append(f"Missing Student IDs: {total_missing_student_id}")
    lines.append(f"Missing required fields (all fields, total issues): {total_missing_fields}")
    lines.append(f"Cross-sheet/duplicate conflicts (differing data on same ID): {len(global_conflicts)}")
    lines.append(f"Over Stay (OS): {total_os}")
    lines.append("")
    lines.append("-" * 32)
    lines.append("")
    status = "HAS VALIDATION ERRORS" if total_invalid or global_conflicts else "READY FOR REVIEW"
    lines.append(f"STATUS: {status}")

    if global_duplicates:
        lines.append("")
        lines.append("Duplicate Student IDs (manual review required):")
        for sid, group in global_duplicates.items():
            sheets_involved = sorted({v.record.source_sheet for v in group})
            conflict_flag = " [CONFLICT]" if sid in global_conflicts else ""
            lines.append(f"  {sid} - appears in {sheets_involved} ({len(group)}x){conflict_flag}")

    report_text = "\n".join(lines)
    print(report_text)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    report_path = OUTPUT_DIR / "validation_report.txt"
    report_path.write_text(report_text, encoding="utf-8")
    print()
    print(f"[written] {report_path}")


def generate_preview(all_validated: list[ValidatedStudent]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    preview_path = OUTPUT_DIR / "normalized_students_preview.csv"

    fieldnames = [
        "student_id",
        "last_name",
        "first_name",
        "middle_name",
        "year_level",
        "course",
        "section",
        "source_sheet",
        "source_row",
        "is_valid",
        "is_over_stay",
        "issues",
    ]

    with preview_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for v in all_validated:
            row = asdict(v.record)
            row["is_valid"] = v.is_valid
            row["is_over_stay"] = v.is_over_stay
            row["issues"] = "; ".join(f"{i.code}: {i.message}" for i in v.issues)
            writer.writerow(row)

    print(f"[written] {preview_path}")


def print_terminal_sample(all_validated: list[ValidatedStudent]) -> None:
    """Never dump the full dataset - show at most one redacted sample."""
    print()
    print("Sample normalized record (redacted):")
    sample = next((v for v in all_validated if v.is_valid), None) or (
        all_validated[0] if all_validated else None
    )
    if not sample:
        print("  (no records extracted)")
        return
    print("  Student ID: [REDACTED]")
    print("  Last Name: [REDACTED]")
    print("  First Name: [REDACTED]")
    print(f"  Year: {sample.record.year_level}")
    print(f"  Course: {sample.record.course}")
    print(f"  Section: {sample.record.section}")
    print(f"  Source sheet: {sample.record.source_sheet}")


if __name__ == "__main__":
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    run(path)
